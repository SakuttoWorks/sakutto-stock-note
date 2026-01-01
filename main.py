import csv
import ctypes
import os
import threading
import tkinter as tk
import traceback
from datetime import date, datetime
from tkinter import messagebox, ttk
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


# ==========================================
#  設定・定数管理
# ==========================================
class Config:
    """アプリケーションの定数設定"""

    OUTPUT_DIR = "【出力結果】株価データ"
    DATA_DIR = "【設定】銘柄リスト"
    EXCEL_FILENAME = "株価一覧.xlsx"
    JP_STOCKS_FILE = "jp_stocks.csv"
    README_SHEET_NAME = "使い方・ABOUT"
    ICON_FILE = "icon.ico"

    # UI設定
    FONT_MAIN = ("Yu Gothic UI", 15)
    FONT_BOLD = ("Yu Gothic UI", 15, "bold")
    FONT_BUTTON_BOLD = ("Yu Gothic UI", 18, "bold")
    FONT_BUTTON_ACCENT = ("Yu Gothic UI", 13)
    FONT_LOG = ("Consolas", 12)
    WINDOW_SIZE = "1000x900+50+50"
    MIN_SIZE = (800, 700)


# ==========================================
#  高DPI対応 (4Kモニタ対策)
# ==========================================
def enable_high_dpi_awareness():
    """Windowsの高DPI設定を有効化"""
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass


# ==========================================
#  データ処理ロジック (Backend)
# ==========================================
class DataManager:
    """ファイル操作やデータ取得に関するロジックを集約"""

    @staticmethod
    def ensure_directories():
        os.makedirs(Config.OUTPUT_DIR, exist_ok=True)
        os.makedirs(Config.DATA_DIR, exist_ok=True)

    @staticmethod
    def load_jp_stock_master() -> Dict[str, str]:
        """銘柄定義ファイルを読み込む。なければデフォルト作成。"""
        master_path = os.path.join(Config.DATA_DIR, Config.JP_STOCKS_FILE)

        # デフォルト作成
        if not os.path.exists(master_path):
            try:
                with open(master_path, "w", encoding="utf-8-sig") as f:
                    f.write(
                        "ticker,name_ja\n2559.T,MAXIS 全世界株式\n7203.T,トヨタ自動車\n"
                    )
                return {"2559.T": "MAXIS 全世界株式", "7203.T": "トヨタ自動車"}
            except Exception:
                return {}

        # 読み込み（エンコーディング自動判別）
        for enc in ["utf-8-sig", "cp932"]:
            try:
                with open(master_path, "r", encoding=enc) as f:
                    reader = csv.DictReader(f)
                    mapping = {}
                    for row in reader:
                        t, n = row.get("ticker"), row.get("name_ja")
                        if t and n:
                            mapping[t.strip()] = n.strip()
                    return mapping
            except (UnicodeDecodeError, Exception):
                continue
        return {}

    @staticmethod
    def fetch_stock_data(
        ticker: str, mode: int, date_range: Optional[Tuple[date, date]]
    ) -> pd.DataFrame:
        """Yahoo Financeから株価データを取得"""
        try:
            tkr = yf.Ticker(ticker)
            kwargs = {"interval": "1d", "auto_adjust": False}

            if mode == 4 and date_range:
                kwargs["start"] = date_range[0].strftime("%Y-%m-%d")
                kwargs["end"] = date_range[1].strftime("%Y-%m-%d")
            else:
                period_map = {1: "1d", 2: "1mo", 3: "3mo"}
                kwargs["period"] = period_map.get(mode, "1d")

            df = tkr.history(**kwargs)
            if df.empty:
                return pd.DataFrame()

            if df.index.tz is not None:
                df.index = df.index.tz_localize(None)

            # 移動平均の計算
            df["MA5"] = df["Close"].rolling(window=5).mean()
            df["MA25"] = df["Close"].rolling(window=25).mean()
            return df
        except Exception:
            return pd.DataFrame()


# ==========================================
#  Excel操作ロジック (Excel Handler)
# ==========================================
class ExcelHandler:
    """OpenPyXLを使用したExcel操作を集約"""

    @staticmethod
    def build_sheet_name(ticker: str, name_ja: str) -> str:
        """Excelシート名に使用できない文字を除外して生成"""
        base = f"{ticker}_{name_ja}" if name_ja else ticker
        for bad in (":", "\\", "/", "?", "*", "[", "]"):
            base = base.replace(bad, "_")
        return base[:31]  # Excelの制限

    @staticmethod
    def create_readme_sheet(wb: Workbook):
        """使い方・免責事項シートの作成"""
        if Config.README_SHEET_NAME in wb.sheetnames:
            ws = wb[Config.README_SHEET_NAME]
        else:
            ws = wb.create_sheet(title=Config.README_SHEET_NAME, index=0)

        ws.sheet_properties.tabColor = "FF9900"

        # コンテンツ定義
        info_text = [
            ("【株価自動追記ツール】", "Title"),
            ("", ""),
            ("■ このファイルについて", "Header"),
            (
                "指定した銘柄の株価（終値・始値・高値・安値）および移動平均線を自動で記録するファイルです。",
                "Normal",
            ),
            (
                "Pythonスクリプトによって、Yahoo! Financeから取得したデータが追記されます。",
                "Normal",
            ),
            ("", ""),
            ("■ 各項目の意味", "Header"),
            ("・日付： データの基準日", "Normal"),
            ("・移動平均： 過去5営業日、25営業日の終値平均", "Normal"),
            ("", ""),
            ("■ グラフの見方", "Header"),
            ("・実線： 終値の推移", "Normal"),
            ("・線（赤/青など）： 各移動平均線", "Normal"),
            ("", ""),
            ("■ 免責事項", "Header"),
            (
                "・本データは Yahoo! Finance (yfinance) の非公式APIを利用しています。",
                "Normal",
            ),
            (
                "・商用利用や、本データに基づく投資判断による損失について、作成者は責任を負いません。",
                "Alert",
            ),
            ("", ""),
            (f"最終更新日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", "Small"),
        ]

        # スタイル定義
        fonts = {
            "Title": Font(size=14, bold=True, color="000080"),
            "Header": Font(size=11, bold=True, color="000000"),
            "Normal": Font(size=10, color="333333"),
            "Alert": Font(size=10, bold=True, color="FF0000"),
            "Small": Font(size=9, color="888888"),
        }
        header_fill = PatternFill(
            start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
        )

        for i, (text, style_key) in enumerate(info_text, start=1):
            cell = ws.cell(row=i, column=2)
            cell.value = text
            cell.font = fonts.get(style_key, fonts["Normal"])
            if style_key == "Header":
                cell.fill = header_fill

        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 80

    @staticmethod
    def setup_worksheet_header(ws: Worksheet, ticker: str, name_ja: str):
        """シートヘッダーの初期設定（A1タイトル、3行目テーブルヘッダー）"""
        if ws["A3"].value is not None:
            return

        # タイトル
        ws["A1"] = f"{ticker} : {name_ja}"
        ws["A1"].font = Font(size=14, bold=True, color="000080")
        ws.merge_cells("A1:D1")

        # テーブルヘッダー
        headers = [
            "日付",
            "終値",
            "始値",
            "高値",
            "安値",
            "5日移動平均",
            "25日移動平均",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="808080", end_color="808080", fill_type="solid"
            )

        # 列幅設定
        widths = {"A": 15, "B": 12, "C": 12, "D": 12, "E": 12, "F": 14, "G": 14}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    @staticmethod
    def merge_and_write_data(ws: Worksheet, df: pd.DataFrame):
        """既存データを読み込み、新規データをマージして書き込む"""
        # 1. 既存データ読み込み
        data_map = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            raw_val = row[0]
            if not raw_val:
                continue

            d_key = None
            if isinstance(raw_val, datetime):
                d_key = raw_val.date()
            elif isinstance(raw_val, date):
                d_key = raw_val
            elif isinstance(raw_val, str):
                try:
                    d_key = datetime.strptime(
                        raw_val.replace("/", "-").split(" ")[0], "%Y-%m-%d"
                    ).date()
                except:
                    pass

            if d_key:
                row_list = list(row)
                row_list[0] = d_key
                data_map[d_key] = row_list

        # 2. マージ処理
        for dt_idx, row in df.iterrows():
            d = dt_idx.date()
            new_row = [
                d,
                float(row["Close"]),
                float(row["Open"]),
                float(row["High"]),
                float(row["Low"]),
                row["MA5"],
                row["MA25"],
            ]
            data_map[d] = new_row  # 上書きまたは新規追加

        # 3. ソートと書き込み
        sorted_dates = sorted(data_map.keys())
        current_row = 4
        number_format = "#,##0"

        for d in sorted_dates:
            values = data_map[d]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                if col_idx > 1 and isinstance(val, (int, float)):
                    cell.number_format = number_format
            current_row += 1

        # 余剰行削除
        if ws.max_row >= current_row:
            ws.delete_rows(current_row, ws.max_row - current_row + 1)

    @staticmethod
    def ensure_table_and_chart(ws: Worksheet, sheet_name: str):
        """テーブル化とグラフの更新"""
        max_row = ws.max_row
        if max_row < 3:
            return

        # テーブル設定
        safe_name = sheet_name.replace(" ", "_").replace(".", "_").replace("-", "_")
        if safe_name[0].isdigit():
            safe_name = "tbl_" + safe_name
        table_name = safe_name[:255]

        table_ref = f"A3:G{max_row}"

        # 既存テーブル確認
        found_tbl = next(
            (tbl for tbl in ws.tables.values() if tbl.displayName == table_name), None
        )
        if found_tbl:
            found_tbl.ref = table_ref
        else:
            new_table = Table(displayName=table_name, ref=table_ref)
            new_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(new_table)

        # グラフ更新
        chart_title = f"{sheet_name} 株価推移"
        # 既存グラフ削除（同じタイトルのもの）
        ws._charts = [c for c in ws._charts if c.title != chart_title]

        chart = LineChart()
        chart.title = chart_title
        chart.style = 2
        chart.y_axis.title = "株価"
        chart.x_axis.title = "日付"
        chart.height = 10
        chart.width = 20

        data_c = Reference(ws, min_col=2, min_row=3, max_row=max_row)
        data_ma5 = Reference(ws, min_col=6, min_row=3, max_row=max_row)
        data_ma25 = Reference(ws, min_col=7, min_row=3, max_row=max_row)
        cats = Reference(ws, min_col=1, min_row=4, max_row=max_row)

        chart.add_data(data_c, titles_from_data=True)
        chart.add_data(data_ma5, titles_from_data=True)
        chart.add_data(data_ma25, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "I3")


# ==========================================
#  UIクラス（Frontend / Controller）
# ==========================================
class StockLoggerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("サクッと！株価ノート v1.0")

        self._setup_window()
        self._init_variables()
        self._create_widgets()

        self.log_safe("ツールを起動しました。更新したい銘柄を選択してください。")

    def _setup_window(self):
        """ウィンドウ初期設定"""
        if os.path.exists(Config.ICON_FILE):
            try:
                self.root.iconbitmap(Config.ICON_FILE)
            except:
                pass

        enable_high_dpi_awareness()
        self.root.option_add("*Font", Config.FONT_MAIN)

        # スタイル設定
        self.style = ttk.Style()
        self.style.configure(".", font=Config.FONT_MAIN)
        self.style.configure("TLabelframe.Label", font=Config.FONT_BOLD)
        self.style.configure("Bold.TButton", font=Config.FONT_BUTTON_BOLD)
        self.style.configure("Accent.TButton", font=Config.FONT_BUTTON_ACCENT)

        self.root.geometry(Config.WINDOW_SIZE)
        self.root.minsize(*Config.MIN_SIZE)

    def _init_variables(self):
        """Tkinter変数の初期化"""
        self.period_var = tk.IntVar(value=1)
        self.start_date_var = tk.StringVar()
        self.end_date_var = tk.StringVar()
        self.progress_var = tk.DoubleVar(value=0.0)
        self.status_msg = tk.StringVar(value="準備完了")

        self.name_map = DataManager.load_jp_stock_master()
        self.ticker_list = list(self.name_map.keys())
        self.check_vars = {}

    def _create_widgets(self):
        """ウィジェットの配置"""
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill="both", expand=True)

        # --- ① 銘柄選択エリア ---
        self._create_stock_selection_area(main_frame)

        # --- ② 期間選択エリア ---
        self._create_period_selection_area(main_frame)

        # --- ③ 実行・進捗エリア ---
        self._create_action_area(main_frame)

        # --- ④ ログエリア ---
        self._create_log_area(main_frame)

    def _create_stock_selection_area(self, parent):
        frame = ttk.LabelFrame(parent, text="① 対象銘柄の選択", padding=15)
        frame.pack(side="top", fill="both", expand=True, pady=10)

        # ボタン群
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(side="top", fill="x", pady=(0, 10))
        ttk.Button(
            btn_frame,
            text="全てチェック",
            command=self.select_all,
            style="Accent.TButton",
        ).pack(side="left", padx=5)
        ttk.Button(
            btn_frame,
            text="チェックを外す",
            command=self.deselect_all,
            style="Accent.TButton",
        ).pack(side="left", padx=5)

        # スクロール付リスト
        canvas = tk.Canvas(frame, borderwidth=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # マウスホイール連動
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", _on_mousewheel)
        self.scrollable_frame.bind("<MouseWheel>", _on_mousewheel)

        if not self.ticker_list:
            ttk.Label(
                self.scrollable_frame,
                text="銘柄データ(jp_stocks.csv)がありません。",
                foreground="red",
            ).pack(pady=10)
        else:
            for t in self.ticker_list:
                name = self.name_map[t]
                var = tk.BooleanVar(value=True)
                self.check_vars[t] = var
                chk = ttk.Checkbutton(
                    self.scrollable_frame, text=f"{t} : {name}", variable=var
                )
                chk.pack(anchor="w", padx=5, pady=5)
                chk.bind("<MouseWheel>", _on_mousewheel)

    def _create_period_selection_area(self, parent):
        frame = ttk.LabelFrame(parent, text="② データ取得期間", padding=15)
        frame.pack(side="top", fill="x", pady=10)

        p_inner = ttk.Frame(frame)
        p_inner.pack(anchor="w")

        modes = [(1, "最新のみ(1日)"), (2, "過去1ヶ月"), (3, "過去3ヶ月")]
        for val, label in modes:
            ttk.Radiobutton(
                p_inner,
                text=label,
                variable=self.period_var,
                value=val,
                command=self.toggle_date_entry,
            ).pack(side="left", padx=10)

        # カスタム期間
        frame_custom = ttk.Frame(frame)
        frame_custom.pack(anchor="w", pady=(15, 0))
        ttk.Radiobutton(
            frame_custom,
            text="期間指定:",
            variable=self.period_var,
            value=4,
            command=self.toggle_date_entry,
        ).pack(side="left", padx=5)

        self.entry_start = ttk.Entry(
            frame_custom,
            textvariable=self.start_date_var,
            width=12,
            state="disabled",
            font=("Consolas", 15),
        )
        self.entry_start.pack(side="left", padx=5)
        ttk.Label(frame_custom, text="〜").pack(side="left")
        self.entry_end = ttk.Entry(
            frame_custom,
            textvariable=self.end_date_var,
            width=12,
            state="disabled",
            font=("Consolas", 15),
        )
        self.entry_end.pack(side="left", padx=5)
        ttk.Label(
            frame_custom,
            text="(YYYY-MM-DD)",
            foreground="#666",
            font=("Yu Gothic UI", 11),
        ).pack(side="left", padx=5)

    def _create_action_area(self, parent):
        container = ttk.Frame(parent)
        container.pack(side="bottom", fill="x", pady=5)

        self.progressbar = ttk.Progressbar(
            container, variable=self.progress_var, maximum=100, mode="determinate"
        )
        self.progressbar.pack(fill="x", pady=(10, 5))

        self.lbl_status = ttk.Label(
            container, textvariable=self.status_msg, foreground="#005500"
        )
        self.lbl_status.pack(anchor="e", pady=(0, 10))

        self.btn_run = ttk.Button(
            container,
            text="株価データの取得・更新を実行",
            command=self.start_processing,
            style="Bold.TButton",
        )
        self.btn_run.pack(fill="x", ipady=15)

    def _create_log_area(self, parent):
        frame = ttk.LabelFrame(parent, text="処理ログ", padding=10)
        frame.pack(side="bottom", fill="x", pady=10)

        self.log_text = tk.Text(frame, height=5, state="normal", font=Config.FONT_LOG)
        self.log_text.pack(side="left", fill="both", expand=True)

        sb = ttk.Scrollbar(frame, orient="vertical", command=self.log_text.yview)
        sb.pack(side="right", fill="y")
        self.log_text["yscrollcommand"] = sb.set

    # --- イベントハンドラ ---
    def select_all(self):
        for var in self.check_vars.values():
            var.set(True)

    def deselect_all(self):
        for var in self.check_vars.values():
            var.set(False)

    def toggle_date_entry(self):
        is_custom = self.period_var.get() == 4
        state = "normal" if is_custom else "disabled"
        self.entry_start.config(state=state)
        self.entry_end.config(state=state)

        if is_custom:
            today_str = date.today().strftime("%Y-%m-%d")
            if not self.start_date_var.get():
                self.start_date_var.set(today_str)
            if not self.end_date_var.get():
                self.end_date_var.set(today_str)

    # --- UI更新ヘルパー ---
    def log_safe(self, msg):
        def _update():
            self.log_text.insert(tk.END, f"{msg}\n")
            self.log_text.see(tk.END)

        self.root.after(0, _update)

    def status_safe(self, msg, progress_val=None):
        def _update():
            self.status_msg.set(msg)
            if progress_val is not None:
                self.progress_var.set(progress_val)

        self.root.after(0, _update)

    # --- メイン処理ロジック ---
    def start_processing(self):
        self.btn_run.config(state="disabled")
        self.progress_var.set(0)
        self.log_safe("=== 処理を開始します ===")
        thread = threading.Thread(target=self.run_logic, daemon=True)
        thread.start()

    def run_logic(self):
        try:
            DataManager.ensure_directories()
            excel_path = os.path.join(Config.OUTPUT_DIR, Config.EXCEL_FILENAME)

            # ターゲット取得
            target_tickers = [t for t, var in self.check_vars.items() if var.get()]
            if not target_tickers:
                self.log_safe("⚠ 銘柄が選択されていません。")
                self.status_safe("待機中 - 銘柄を選択してください", 0)
                self.finalize_ui()
                return

            # 日付範囲設定
            mode = self.period_var.get()
            d_range = None
            if mode == 4:
                try:
                    s = datetime.strptime(self.start_date_var.get(), "%Y-%m-%d").date()
                    e = datetime.strptime(self.end_date_var.get(), "%Y-%m-%d").date()
                    d_range = (s, e)
                except ValueError:
                    self.log_safe(
                        "⚠ 日付形式が不正です。デフォルト(最新)で実行します。"
                    )
                    mode = 1

            # Excelオープン
            self.status_safe("Excelファイルを開いています...")
            wb = self.open_workbook_safely(excel_path)
            if not wb:
                self.finalize_ui()
                return

            ExcelHandler.create_readme_sheet(wb)
            for trash in ["Sheet", "Sheet1"]:
                if trash in wb.sheetnames and len(wb.sheetnames) > 1:
                    del wb[trash]

            total = len(target_tickers)
            self.log_safe(f"▶ 対象: {total} 銘柄")

            for i, ticker in enumerate(target_tickers):
                progress = (i / total) * 100
                self.status_safe(f"データ取得中... ({i+1}/{total}): {ticker}", progress)

                df = DataManager.fetch_stock_data(ticker, mode, d_range)
                if df.empty:
                    self.log_safe(f"  [Skip] {ticker}: データなし")
                    continue

                name_ja = self.name_map.get(ticker, "")
                sheet_name = ExcelHandler.build_sheet_name(ticker, name_ja)
                self.log_safe(f"  [OK] {ticker} -> {sheet_name}")

                # シート取得または作成
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                elif ticker in wb.sheetnames:
                    ws = wb[ticker]
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)

                ExcelHandler.setup_worksheet_header(ws, ticker, name_ja)
                ExcelHandler.merge_and_write_data(ws, df)
                ExcelHandler.ensure_table_and_chart(ws, sheet_name)

            self.status_safe("Excelファイルを保存しています...", 95)
            if self.save_workbook_safely(wb, excel_path):
                self.status_safe("完了しました", 100)
                self.log_safe(f"✔ 全ての処理が完了しました: {excel_path}")
            else:
                self.status_safe("保存キャンセル", 0)

        except Exception as e:
            self.log_safe(f"⚠ エラー発生: {e}")
            traceback.print_exc()
        finally:
            self.finalize_ui()

    def finalize_ui(self):
        self.root.after(0, lambda: self.btn_run.config(state="normal"))

    def open_workbook_safely(self, path):
        if not os.path.exists(path):
            return Workbook()
        while True:
            try:
                return load_workbook(path)
            except PermissionError:
                if not messagebox.askretrycancel(
                    "ファイルが開かれています",
                    f"Excelファイル '{os.path.basename(path)}' を閉じてから「再試行」を押してください。",
                ):
                    self.log_safe("キャンセルされました。")
                    return None
            except Exception as e:
                self.log_safe(f"ファイル読み込みエラー: {e}")
                return Workbook()

    def save_workbook_safely(self, wb, path):
        while True:
            try:
                wb.save(path)
                messagebox.showinfo(
                    "完了",
                    "データの更新が完了しました！\nExcelファイルを開いて確認してください。",
                )
                return True
            except PermissionError:
                if not messagebox.askretrycancel(
                    "保存できません",
                    "Excelファイルが開かれているため保存できません。\n閉じてから「再試行」を押してください。",
                ):
                    self.log_safe("保存をキャンセルしました。")
                    return False


if __name__ == "__main__":
    root = tk.Tk()
    app = StockLoggerApp(root)
    root.mainloop()
